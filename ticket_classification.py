# -*- coding: utf-8 -*-
"""
Created on Thu Feb  1 08:04:24 2018
USE PYTHON 3
More info on how to use it in Readme.txt
@author: cz2fzk003
"""

# -*- coding: utf-8 -*-
"""
Created on Mon Jan 15 14:31:47 2018

@author: faridull, cz2fzk003
"""

import os, pandas as pd, xlrd, numpy as np,  matplotlib.pyplot as plt, \
time, math, mpld3, logging, fnmatch
from collections import defaultdict 
import  string, re
#from nltk.corpus import stopwords
from openpyxl import load_workbook
from mpld3 import plugins
# adjust pandas parameters for html plots later
pd.options.display.width = 150
pd.options.display.max_colwidth = 1000 # for finding computer name within Beschreibung, we need to increase
# that value quite a bit, will be changed in html code to 150 

print("Job was started successfully. Please hold on while the job is running")
# <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< #
## DATA PREPROCESSING ################
# <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< #
# set path
# for testing only comment out following two lines for pyinstaller
# path = 'C:\\1.DataScience_Projects\\d.IT_HD'
# os.chdir(path)

# Directory of .exe file # for later
exe_dir = os.getcwd()
os.chdir(exe_dir)
# set up log file (will be in working directory of exe file)

logging.basicConfig(filename='logfile.txt', level=logging.INFO, 
                    format='%(asctime)s %(levelname)s %(name)s %(message)s')
logger = logging.getLogger(__name__)

# <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< #
## SUPPORT FILE ################
# <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< #
def load_support_file(filename):
    try: 
        book = xlrd.open_workbook(os.path.join("Files/", filename), formatting_info=True)
        sheet = book.sheet_by_index(0)
        original_df = pd.read_excel(os.path.join("Files/",filename), sheet_name = "Sheet1")
        
        # get color code of all cells
        rows, cols = sheet.nrows,sheet.ncols
        color_code_df = pd.DataFrame()
        
        for row in range(rows):
            #print("row:",  row)
            for col in range(cols):
                #print("col:", col)
                #name_of_cell = sheet.cell(row,col).value
                color_of_cell_helper = sheet.cell_xf_index(row,col)
                color_of_cell = book.xf_list[color_of_cell_helper].background.pattern_colour_index
                #print(color_of_cell)
                color_code_df.loc[row, col] = color_of_cell
        
        # replace colors with weights
        # zero = no color (empty cell), green = weight 1, yellow = weight 2, red = weight 3
        vals_to_replace = {64:0, 50:1, 13:2, 10:3}
        color_code_df = color_code_df.replace(vals_to_replace)
        # remove first col and row 
        color_code_key = color_code_df.iloc[1:, 1:]
        # as a list of lists
        color_list = color_code_key.values.tolist()
        # as a pure list (Farid liked lists a lot...)
        color_list1 = sum(color_list, [])
        
        # only where >0 (zeros are blank cells, not filled with any colors.)
        color_list1_nonan = [x for x in color_list1 if x > 0]
        
        #filtering keywords from support file
        df_filter_keywords = original_df.iloc[:,1:]
        #list of keywords (list of lists)
        dfList2 = df_filter_keywords.values.tolist()
        #color_code_df.iloc[0,1]
        # as a list
        dfList = sum(dfList2, [])
        #removing nan (empty cells, non-colored)
        dfList1 = [i for i in dfList if i is not np.nan]
        
        # put into a dictionary (list with different list- lenght was checked, but
        # not the most elegant option here)
        dictionary_keywords = dict(zip(dfList1, color_list1_nonan))
        
        # the original file has several words for open / closed tickets in a "Status" columns. 
        # For better filtering
        # Therefore, in a new sheet, several words were listed with their respective mapping
        # mapping. This sheet can be modified by their team. 
        # The mapping will be joined with the original file later
        open_closed_mapping = pd.read_excel(os.path.join("Files/", filename), sheet_name = 'Mapping_open_closed')
        
        # log everything!
        text = "Support file was read and pre-processed."
        print(text)
        logger.info(text)
        return (dictionary_keywords, original_df, open_closed_mapping)
    except Exception as e: 
        print("Support file cannot be read / found - please check")
        # log error
        logger.error(e)
        
# function call
dictionary_keywords, original_df, open_closed_mapping = load_support_file(filename = "supporting_info_aw.xls")
 
  
# <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< #
## PREPROCESS HELPDESK (HD) FILE #############
# <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< #
def read_and_preprocess_HD_tickets():
    try: 
        # read xlsx file (actual ticket file). Filename changes, therefore we only know it is
        # an .xlsx file. If two or more files are detected, the user is informed via logging and the
        # first file is read. In the readme, it is specified that only one .xlsx file should be in the
        # respective folder.
        pattern = '*.xlsx'
        file_to_read = fnmatch.filter(os.listdir("Files/"), pattern)
        if len(file_to_read)> 1:
            text = "More than one .xlsx file was detected. Please check,\
            there should only be one file. The results may not be correct now."
            print(text)
            logger.info(text)
        df = pd.read_excel(os.path.join("Files/", file_to_read[0]), index = False)

        # The mapping file was read in in the previous step  and is then left-joined
        # to the original file. 
        df = pd.merge(df, open_closed_mapping, how = "left", on = ["Status"] )
        
        # Geöffnet or Opened (depending on language, we need to read in different columns)
        # then, month and week columns are added 
        if("Geöffnet" in df.columns) == True:
            col_name_opened = "Geöffnet"
            text_lang = "Detected GSN column header language is German"
            print(text_lang)
            logger.info(text_lang)
        else: 
            col_name_opened = "Opened"
            text_lang = "Detected GSN column header language is English"
            print(text_lang)
            logger.info(text_lang)
        pd.to_datetime(df[col_name_opened])
        
        df['year'] = pd.DatetimeIndex(df[col_name_opened]).year 
        df['month'] = pd.DatetimeIndex(df[col_name_opened]).month
        df['week'] = pd.DatetimeIndex(df[col_name_opened]).week
       # df["year_week"] = df['year'].map(str) + "_" + df['week'].map(str)
        df['year_week'] = df[col_name_opened].dt.strftime('%Y-%W')
        # pd.to_datetime(df[col_name_opened])
        # week 52 remove - non complete, only for testing purposes
        # df = df[(df.week != 52) & (df.year == 2017)]

        # extract text columns from problem and solution columns - 
        # after language detection (word Beschreibung is only in the German version)
        if("Beschreibung" in df.columns) == True:
            cols_of_interest_TM = ['Beschreibung', 'Kurzbeschreibung','Auflösungshinweise','year_week', "year", "week"]
        else: 
            cols_of_interest_TM = ['Description', 'Short Description','Solution','year_week', "year", "week"]
            
        text_df = df[cols_of_interest_TM]
    
        # join text df columns for text mining
        text_all = ''.join(text_df.to_string(header = False))
        
        # text_all = ''.join(df.to_string(header = False))
        text = text_all.lower().split("\n")
        
        ## find keywords in text and add as columns 
        # keywords have different weights which need to be added.
        for word, value in dictionary_keywords.items():
            new_word = word.lower()
            new_column = []
            for line in text:
                if new_word in line:
                    #add condition for colors
                    new_column.append(value)
                else:
                    new_column.append(0)
            # put into new column
            df[word] = pd.Series(new_column)
        
        # various keywords belong to one topic. topics are in rows, with respective
        # keywords in the same row
        # therefore, we need to sum weighted keywords to topics
        for i in range(original_df.shape[0]):
            topic = original_df.iloc[i,0]
            # print(topic)
            keyword_for_topic = original_df.iloc[i,1:]
            # remove nans
            keyword_for_topic = keyword_for_topic[keyword_for_topic.notnull()]
            # print(keyword_for_topic)
            # detect in original dataframe
            df_only_keyword_columns = df.loc[:,keyword_for_topic]
            # sum up
            sum_only_keyword_columns = df_only_keyword_columns.sum(axis = 1)
            # create new topic column
            df[topic] = sum_only_keyword_columns
       
    #summing keywords to topic NOT WORKING -- THEREFORE VIA DF (list was problematic due to differing lenths)
    #for i in range(len(dfList2)):
    #    col = filter(lambda v: v==v, dfList1[i])
    #    df[original_df.iloc[i][0]] = df[col].sum(axis = 1)
    
        only_topic_cols = df.loc[:, original_df.iloc[:,0]]
        
        # only one topic should be selected for one ticket -- 
        # therefore, select topic with highest value
        # but only if not all are zero 
         # add an "unassigned column" for future processing
        df["Unassigned"] = 0
          
        for i in range(only_topic_cols.shape[0]):
            df.loc[i, "Assigned_topic"] = np.where(only_topic_cols.iloc[i,:].sum()==0, "Unassigned", only_topic_cols.iloc[i,:].idxmax()).tolist()
        
        # put assigned topic as first column
        cols = df.columns.tolist()
        df = df[[cols[-1]] + cols[:-1]]     
            
        # normalizing data - set every topic to 0 except for the assigned topic (which gets a 1)
        for i in range(only_topic_cols.shape[0]):
            only_topic_cols.iloc[i,df.loc[i, "Assigned_topic"] == only_topic_cols.iloc[i,:].index] = 1
            only_topic_cols.iloc[i,df.loc[i, "Assigned_topic"] != only_topic_cols.iloc[i,:].index] = 0
        
        # replace in original df
        df[only_topic_cols.columns] = only_topic_cols
          
        # log everything
        text = "Helpdesk file(s) was/ were read and pre-processed."
        print(text)
        logger.info(text)
        
        return(df,only_topic_cols)
    except Exception as e:
        # log error
        logger.error(e)       
  
# function call
df, only_topic_cols = read_and_preprocess_HD_tickets()

# <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< #
## ADD COLUMN WITH COMPUTER NAME #############
# <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< #
# The computer name (7 / 8 strings, 7 digits, starting with DE)
# is hidden in the Beschreibungs-Column and should be extracted and put into a new column
# set up a re.compile and find a string matching criteria
def Add_computer_name(df): 
    try:
        check = re.compile("DE[A-Z][A-Z][A-Z][A-Z][A-Z][0-9][0-9][0-9][0-9][0-9][0-9][0-9]|DE[A-Z][A-Z][A-Z][A-Z][A-Z][A-Z][0-9][0-9][0-9][0-9][0-9][0-9][0-9]")
        df = df.reset_index() # for looping over df
        # find correct column
        if("Beschreibung" in df.columns) == True:
            col_of_interest = ['Beschreibung']
        else: 
            col_of_interest= ['Description']
        # create new column Computer name based on string check (see above)    
        df["Computer_name"] = ""     
        
       # for i in range(df.shape[0]):
        #    df.loc[i, "Computer_name"] = re.findall(check, df.loc[i, col_of_interest].to_string())
          
        df["Computer_name"] = df[col_of_interest].apply(lambda x: re.findall(check, x.to_string()), axis = 1)
        # put computer name to front and remove column "index"
        del df["index"]
        cols = df.columns.tolist()
        df = df[[cols[-1]] + cols[:-1]]  
        text = "Computer name was detected and added as a new column."
        print(text)
        logger.info(text)
        return(df)
    except Exception as e:
        print("Computer name could not be detected")
        print(e)
        # log Error
        logger.error(e)
        
df = Add_computer_name(df)
# <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< #
## WEEKLY ANALYSIS / PLOTS  ##################
# <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< #  
# weekly and monthly frequency of topics and keywords

def weekly_analysis():
    try: 
        weekly = df.groupby(['year_week']).sum().drop(['month'], axis = 1)
        weekly = weekly.reset_index()
        # create new only topic cols 
        only_topics_cols_incl_unassigned = only_topic_cols.columns.tolist()
        only_topics_cols_incl_unassigned.append("Unassigned")
        
        # extract only topic columns
        weekly_topics = weekly[only_topics_cols_incl_unassigned]
        
        # month to week percentage change with 4 week rolling average
        Monthly_mean = weekly_topics.rolling(4).mean().add_suffix('_4wk_roll_avg')
        
        Monthly_mean.loc[0] = np.nan 
        Monthly_mean.index = Monthly_mean.index + 1  # shifting index for calendar year month
        Monthly_mean = Monthly_mean.sort_index()       
        index = pd.Index(range(len(Monthly_mean)))
        Monthly_mean = pd.DataFrame(Monthly_mean, index=index)              
        # merge via week index
        df_merged = Monthly_mean.merge(weekly_topics, how='right', left_index=True, right_index=True)                
        #df_cat = pd.concat([Monthly_mean, weekly_topics], axis=1)
        
        weekly_avg= df_merged[[item for items in zip(weekly_topics.columns, Monthly_mean.columns) for item in items]]
        weekly_avg = weekly_avg.dropna(axis=0, how='all')
            
        #month to week %  change to dataframe
        data = pd.DataFrame()
        
        for i in range(0, len(weekly_avg.columns), 2):
            data = data.append(((weekly_avg.iloc[:, i] - weekly_avg.iloc[:, i+1])/weekly_avg.iloc[:, i+1]), ignore_index=True)
            data_fin = data.transpose()

        data_fin.columns = only_topics_cols_incl_unassigned
        data_fin = data_fin.round(2)*100
        data_fin['year_week'] = weekly['year_week'] 
        data_fin.set_index('year_week', inplace=True)  
                         
        # week to week % changes 
        #percentage_change_weekly = (weekly_topics.pct_change(axis = 0).round(2)*100)
        weekly2 = df.groupby(['year_week']).sum().drop(['month','week'], axis = 1)       
        # only topics
        weekly_topics2 = weekly2[only_topics_cols_incl_unassigned]
                                 
        # weekly percentage changes 
        percentage_change_weekly = (weekly_topics2.pct_change(axis = 0).round(2)*100)
        
        # percentage change of last week --> needs to be flexible as weeks change!
        last_week_pct = percentage_change_weekly.iloc[-1:]                      
        last_week_freq = weekly_topics2[-1:]                  
        frames = [last_week_freq, last_week_pct ]
        last_weeks = pd.concat(frames, keys = ['Frequency', 'Pct_Change'])
        
        # only last weeks file 
        # last weeks dataframe only
        df2 = df[df.year == df.year.max()]
        df2 = df2[df2.week == df2.week.max()]
    
        #  create a results directory inside the exe folder - if exists, simply write into Results dir
        if not os.path.exists("Results"):
            os.makedirs("Results")
        
        # plotting - either year or last weeks
        weekly_topic_df = [weekly_topics2, weekly_topics2.tail()]
        weekly_topic_df_name = ["weekly_topics_year", "last_months_topics"]
        for i in range(len(weekly_topic_df)):
            print(i)
            plotstr = time.strftime("%Y%m%d_time_%H%M_")
            plotstr = plotstr + weekly_topic_df_name[i] + ".png"
            color_list = plt.cm.Paired(np.linspace(0,1, len(only_topics_cols_incl_unassigned)))
            f = plt.figure() 
            # only integers, no decimal places on x-axis
            #xint = range(min(weekly_topic_df[i].index), math.ceil(max(weekly_topic_df[i].index))+1)
            # adjust axis labels, title and tick marks 
            #plt.xticks(xint, size = 14)                                  
            plt.title('Weekly topic frequencies')
            weekly_topic_df[i].plot(kind='line', ax=f.gca(), figsize = (15, 4), color = color_list)# cmap = "gist_ncar"
            plt.legend(loc='center left', bbox_to_anchor=(1.0, 0.5))
            plt.savefig(os.path.join("Results/", plotstr), bbox_inches='tight')
            #plt.show()
        # log everything
        text = "Weekly analysis completed, plots added to the directory"
        print(text)
        logger.info(text)
        return(weekly_topics, data_fin, df2, last_weeks, percentage_change_weekly, only_topics_cols_incl_unassigned, weekly2, weekly_topics2)
    
    except Exception as e:
        # log error
        logger.error(e) 
        
# function call
weekly_topics, data_fin, df2, last_weeks, percentage_change_weekly, only_topics_cols_incl_unassigned, weekly2, weekly_topics2 = weekly_analysis()



# <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< #
## FINDING NEW FREQUENT WORDS ################
# <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< #
def New_Keywords():
    try:
        stop_words =  open(os.path.join("Files/", "stopwords.txt"),"r")
        
        # extract text columns after language detection
        if("Beschreibung" in df2.columns) == True:
            cols_of_interest_TM = ['Beschreibung', 'Kurzbeschreibung','Auflösungshinweise','year_week', "year", "week"]
        else: 
            cols_of_interest_TM = ['Description', 'Short Description','Solution','year_week', "year", "week"]
            
        text_df = df2[cols_of_interest_TM]
    
        # join text df columns for text mining
        text_df_str = ''.join(text_df.to_string(header = False))
        text_for_freq = text_df_str.lower().split("\n")
        text_for_freq  = ' '.join(text_for_freq)
        #text_for_freq=text_df_str.encode("ascii", "ignore")
        #stop = set(stopwords.words('english'))
        stop = list(stop_words)
        for word in stop:
            stop = [word.strip() for word in stop]
        ##    if word not in stop or len(word)==1:
        #        stop.add(word.strip())
        #stop_words_german = stop_words.get_stop_words('german')
        #stop = stop + stop_words_german
            
        ##removing digits and puntuations
        for char in list(string.digits):
           text_for_freq = text_for_freq.replace(char, "")
        for character in list(string.punctuation):
            text_for_freq = text_for_freq.replace(character, "")
            text_for_freq = text_for_freq.rstrip(string.whitespace).lower()
               
        def word_frequency(text_for_freq, stop):
            wordcount = defaultdict(int)
            for word in text_for_freq.split(): # get rid of punct. marks and stop words.
                if word not in stop and len(word)>=3:
                    wordcount[word] += 1
                    # wordcount.items() # the full list    (tuple(k,v))
            return wordcount.items()
        
        ##finding frequency
        word_freq = list(word_frequency(text_for_freq, stop))
        #word_freq = list(sorted(word_freq, reverse = True))
        word_freq.sort(key=lambda x: x[1], reverse = True)
        # keywords file - in the new word list, we should not have keywords        
        key = [i.lower() for i in dictionary_keywords.keys()]
        # only top new words
        words_found = [i for i in word_freq[:80]]
        # getting rid of "new" words that are already in keywords list
        new_words_found = [x for x in words_found if x[0] not in key]
        # create a new df
        News = pd.DataFrame([x for x in new_words_found], columns=['New_word', 'Frequency'])
        last_week_year = text_df.year_week.iloc[1]
        News['Year_Week'] = last_week_year
        
        # log everything
        text = "New keywords were detected."
        print(text)
        logger.info(text)
        
        return (News, last_week_year)
    except Exception as e:
        print("New keywords cannot be generated.")
        # log Error
        logger.error(e)
#calling function
News, last_week_year = New_Keywords()  

##<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<##
## PREPARE EXCEL OUTPUT ##########
##<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<##
def excel_output_HD_results(): 
    try:
        #color greater than 20 % change
        def color_twenty_red(val):
            color = 'red' if val > 20 else 'black'
            return 'color: %s' % color
        
        # two different versions of highlighting changes was needed for the file
        # were we have absolute values, but not % changes
        def highlight_change(data, color='red'):
            '''
            highlight the percentage change more than 20% in a Series or DataFrame
            '''
            attr = 'background-color: {}'.format(color)
            if data.ndim == 1:  # Series from .apply(axis=0) or axis=1
                is_red = data.pct_change()
                return [attr if v > 0.20 else '' for v in is_red]
            else:  # from .apply(axis=None)
                is_red = data.pct_change() > 0.20
                return pd.DataFrame(np.where(is_red, attr, ''),
                                    index=data.index, columns=data.columns)
        # excel_file_string with date and time for master file 
        
        timestr = os.path.join("Results/", time.strftime("%Y%m%d_time_%H%M_Results_File.xlsx"))
        #book = load_workbook(timestr)
        writer2 = pd.ExcelWriter(timestr, engine='xlsxwriter')
        #writer = pd.ExcelWriter(timestr, engine='openpyxl') 
        #writer.book = book
        #writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df.to_excel(writer2, sheet_name = "Main_file",  index = False)
        df2.to_excel(writer2, sheet_name = "Main_file_last_week_only",  index = False)
        data_fin.to_excel(writer2, sheet_name = "Week_to_month_pct_change")
        weekly_topics2.to_excel(writer2, sheet_name = "Weekly_Topics_abs_freq")
        percentage_change_weekly.to_excel(writer2, sheet_name = "Weekly_pct_change")

        # style.format("{:.2%}").apply(highlight_change, color='red', axis=None).\
                             
        last_weeks.T.to_excel(writer2, sheet_name = "Last_week_change") 
        # drop NaN from new word list
        News.dropna(subset = ['New_word'], inplace  = True)
        News.to_excel(writer2, sheet_name = "New Keywords",  index = False)                     
        
        number_rows = len(percentage_change_weekly)

        # Get access to the workbook and sheet
        workbook = writer2.book
        worksheet_weekly = writer2.sheets['Weekly_pct_change']
        worksheet_monthly = writer2.sheets['Week_to_month_pct_change']
        

        # Define our range for the color formatting
        color_range = "B2:Z{}".format(number_rows+1)
        
        # Add a format. White fill with  red text.
        format1 = workbook.add_format({'bg_color': '#FFFFFF',
                                       'font_color': '#FF0000'}) 
        # Highlight the >20 values in Red
        worksheet_weekly.conditional_format(color_range, {'type':     'cell',
                                                   'criteria': '>',
                                                   'value':    20,
                                                   'format':   format1})
    
        worksheet_monthly.conditional_format(color_range, {'type':     'cell',
                                                   'criteria': '>',
                                                   'value':    20,
                                                   'format':   format1})
    
        writer2.save()
        
        # log everything
        text = "Excel output files were created successfully and saved in Results folder"
        print(text)
        logger.info(text)
        
    except Exception as e:
    # log error
        logger.error(e) 
    
# function call    
excel_output_HD_results()

#############################################
### HTML PLOT ###############################
#############################################
# check out https://mpld3.github.io/examples/scatter_tooltip.html
def preprocess_for_html_output():
    try: 
        # get max of year and then max week for last weeks info only.
        df2 = df[df.year == df.year.max()]
        df2 = df2[df2.week == df2.week.max()]
        df2 = df2.reset_index()
        pd.options.display.max_colwidth = 150 # it was larger for finding computer name in Beschreibung
        for i in range(df2.shape[0]):
            # get all topics
            topics =  pd.Series(only_topics_cols_incl_unassigned)
            # select the major topic of row
            row_with_topic = original_df[df2.loc[i, "Assigned_topic"] == topics]
            # only works if row not empty
            if len(row_with_topic) > 0: 
                # print(row_with_topic.iloc[0,0])
                # find keywords for that topic
                keyword_for_topic = row_with_topic.iloc[0,1:].tolist()
                #removing nan
                keyword_for_topic = [f for f in keyword_for_topic if f is not np.nan]
                a = df2.loc[i,keyword_for_topic]
                df2.loc[i, "Max_Keyword"] =  a.loc[a == a.max()].index[0]
            else:
                df2.loc[i, "Max_Keyword"] = "Unassigned"      
                          
        
        # create a grouped by dataframe to only have unique keywords and count occurence of keywords
        df_grouped = df2.groupby(["Max_Keyword", "Assigned_topic"]).count().reset_index().iloc[:,0:3]
        df_grouped = df_grouped.rename(columns={"index": "keyword_freq", "Max_Keyword": "keywords", "Assigned_topic":"topics"})
        
        # remove unassigned 
        #df_grouped = df_grouped[df_grouped['keywords'] != "Unassigned"]
        #df_grouped = df_grouped.reset_index()
        # convert topics to categorical and create category codes for scatter plotting 
        df_grouped.topics = pd.Categorical(df_grouped.topics)
        df_grouped.keywords = df_grouped.keywords.astype("category")
        df_grouped['topic_code'] = df_grouped.topics.cat.codes
        
        # log everything
        text = "Data was preprocessed for html output"
        print(text)
        logger.info(text)
        return(df2, df_grouped) 
    except Exception as e:
    # log error
        logger.error(e) 

# function call
df2, df_grouped = preprocess_for_html_output()    
    
# for scatter plotting, some jitter is nice to avoid overlap. the number for stdev can
# be adapted, depending on how far the points should be away from each other.
# function will be called in the plotting phase, instead of calling plt.scatter.
def rand_jitter(arr):
    stdev = .015*(max(arr)-min(arr))
    return arr + np.random.randn(len(arr)) * stdev

def jitter(x, y, s=20, c='b', marker='o', cmap=None, norm=None, vmin=None, vmax=None, alpha=None, linewidths=None, verts=None, hold=None, **kwargs):
    return plt.scatter(rand_jitter(x), rand_jitter(y), s=s, c=c, marker=marker, cmap=cmap, norm=norm, vmin=vmin, vmax=vmax, alpha=alpha, linewidths=linewidths, verts=verts, hold=hold, **kwargs)
  
# this is the main function for creating a browser-based plot. 
# the plot is a jittered scatter plot with the topics on the x-axis
# and the frequency of keywords on the y-axis. The frequency of keywords
# currently also affects the size of the dot (can be changed though of course)
# the mouseover then gives more information on the keyword: either only the keyword
# or also the ticket information texts to get a better understanding of the 
# keyword context.
# the argument "full_info_mouseover" declares whether the mouse-over only includes
# the keyword only or full information (row text)

def create_html_plot(full_info_mouseover = True):
    #fig, ax = plt.subplots(subplot_kw=dict(axisbg='#EEEEEE'))
    try:      
        fig = plt.figure(figsize=(15, 8))
        
        scatter = jitter(df_grouped.topic_code, df_grouped.keyword_freq, s = (df_grouped.keyword_freq*40), \
                         c = df_grouped.topic_code, alpha = 0.7, cmap=plt.cm.get_cmap('prism'))
        # set y-axis to only include integers (no decimal places, doesn't make sense for frequencies)
        # yint = range(min(df_grouped.keyword_freq), math.ceil(max(df_grouped.keyword_freq))+1)
        # adjust axis labels, title and tick marks 
        # for x axis ticks, replace cat codes with category strings
        # plt.xticks(df_grouped.topic_code, df_grouped.topics, size = 12, rotation=90)
        plt.xticks(np.arange(len(tuple(df_grouped.topics.unique()))), sorted(tuple(df_grouped.topics.unique())), size = 11, rotation = 90)
        plt.xlabel("Topics", size = 22)
        plt.ylabel("Frequency of keywords in tickets", size = 22)
        plt.title("Number of tickets for keywords per topic (last week) - mouse over for details", size = 26)
        plt.yticks(size = 20)
        plt.grid(color = "gray", linestyle = "dashed")
        # for simple mouseover (only keywords), this is very easy - just declare the keywords
        if full_info_mouseover == False:
            plotstr2 = time.strftime("%Y%m%d_time_%H%M_")
            plotstr2 = os.path.join("Results/", plotstr2 +  "keyword_interactive_plot.html")
            labels=df_grouped.keywords.tolist()
            tooltip = mpld3.plugins.PointLabelTooltip(scatter, labels=labels)
        # for full mouseover, we need to define some parameters and get the problem
        # text info from the original (non-grouped) dataframe.
        else: 
             # Define some CSS to control our custom labels
            css = """
            table
            {
              border-collapse: collapse;
              width: 100%;
            }
            tr:nth-child(n + 25) {
              visibility: hidden;
            }
            th
            {
              color: #ffffff;
              background-color: #4f2f2f;
            }
            td
            {
              background-color: #ffff99;
              max-width: 550px;
              overflow: hidden;
              text-overflow: ellipsis;
              white-space: nowrap;
            }
            table, th, td
            {
              font-family:Arial, Helvetica, sans-serif;
              font-size:small;
              border: 1px solid black;
              text-align: left;
            }
            """
            # loop through the grouped data frame, get keyword, find that keyword
            # in the original dataframe and get the problem texts for that keyword
            # since the problem text is not included in the grouped dataframe anymore
            #######################
            #CHANGE TEXT COLUMN ##
            ######################
            labels = []
            df2["Max_Keyword"] = " " + df2["Max_Keyword"] + " "
            for i in range(len(df_grouped)):
                # get keyword of row
                keyword = df_grouped.loc[i,"keywords"] 
                # add white spaces for exact matching (to avoid Sap_h in being in Sap_h AND Sap_hana)
                keyword2 = " " + keyword + " "
                # find texts from the first (ungrouped dataframe) matching that keyword
                problem_texts_with_keyword = df2[df2["Max_Keyword"].str.contains(keyword2, case = False)]
                # convert keywords and texts in problem definition to dataframe, rename columns for binding together
                # ATTENTION: NOW Windows 10 and Windows both match sometimes - only first used, search for
                # more elegant way!!!
                keyword_as_df = pd.DataFrame(pd.Series(problem_texts_with_keyword["Max_Keyword"].unique()[0]))
                keyword_as_df = keyword_as_df.rename(columns = {0:"Content"})
                
                # Kurzbeschreibung or short description (language of input)
                if("Kurzbeschreibung" in df2.columns) == True:
                    col_short_description = 'Kurzbeschreibung'
                else: 
                    col_short_description =  'Short Description'
                            
                texts = pd.DataFrame(problem_texts_with_keyword[col_short_description].unique())
                texts = texts.rename(columns = {0: "Content"})
                # bind into one df
                label = keyword_as_df.append(texts)
                # now we also want to have an index which indicates whether the word displayed
                # is a problem description or a keyword --> first should be the keyword, then the texts
                # create a list with the word "text" for problem definition as often as we have rows with a problem
                # with that keyword
                numrows_text = np.repeat("Problem Text", texts.shape[0]).tolist()
                numrows_text.insert(0, "Keyword")
                # put that into a dataframe and re-index (necessary for tooltip)
                label['Type'] = numrows_text
                label.set_index('Type', inplace=True)
                # convert to html object
                labels.append(str(label.to_html()))
            # now create the tooltip object
            tooltip = mpld3.plugins.PointHTMLTooltip(scatter, labels=labels,  css=css)
            plotstr2 = time.strftime("%Y%m%d_time_%H%M_")
            plotstr2 = os.path.join("Results/", plotstr2 +   "keyword_and_text_interactive_plot.html")
        # independent of full or partial mouseover, call d3 --> browser and kernel need restart after calling  show()
        mpld3.plugins.connect(fig, tooltip)
        mpld3.save_html(fig, plotstr2)
        #mpld3.save_html(fig, "test.html")
        # log everything
        text = "Html plots were created and saved in the results folder"
        print(text)
        logger.info(text)
    except Exception as e:
        # log error
        logger.error(e) 
    
# call function
create_html_plot(full_info_mouseover = True)
create_html_plot(full_info_mouseover = False)


# using cmd (commandline) go to the correct directory,
# then type pyinstaller --onefile test_exe.py 
# will create a standalone executable

# <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< #
## OLD: FINDING NEW FREQUENT WORDS ################
# <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< #
'''
text_for_freq=text_all.encode("ascii", "ignore")
stop = set(stopwords.words('english'))
new_words=open("german.txt", "r") #for german stop_words_de
for word in new_words:
    if word not in stop or len(word)==1:
        stop.add(word.strip())
##removing digits and puntuations
for char in list(string.digits):
   text_for_freq = text_for_freq.replace(char, "")
for character in list(string.punctuation):
    text_for_freq = text_for_freq.replace(character, "")
    text_for_freq = text_for_freq.rstrip(string.whitespace).lower()


def word_frequency(text_for_freq, stop):
    wordcount = defaultdict(int)
    for word in text_for_freq.split(): # get rid of punct. marks and stop words.
        if word not in stop and len(word)>=3:
            wordcount[word] += 1
            # wordcount.items() # the full list    (tuple(k,v))
    return wordcount.items()

##finding frequency
word_freq = word_frequency(text_for_freq, stop)

##for plotting
#word_freq.sort(key=lambda x: x[1], reverse=True)
#words, frequency = zip(*word_freq)
#indices = np.arange(len(words))
##print word_freq[0:30]
#
df_keywords = original_df.values.tolist()
df_keywords_list = sum(df_keywords, [])
df_keywords_list_only = [i for i in df_keywords_list if i is not np.nan]
key = [i.lower() for i in df_keywords_list_only]
words_found = [i for i in word_freq[:50]]
#printing words that are not in keywords
new_words_found = [x for x in words_found if x[0] not in key]
#
News = pd.DataFrame([x for x in new_words_found], columns=['New_word', 'Frequency'])'''
#writer = pd.ExcelWriter('C:\\Users\\faridull\\Desktop\\project IT_Tickets\\try.xlsx', engine='xlsxwriter')
#normalize_topics.to_excel(writer,'topic_frequency', index = False)
#df_zeros.to_excel(writer,'uncatogorize_tickets')
#News.to_excel(writer,'New_words', index = False)
#last_weeks.to_excel(writer,'last_week', index = False)
#writer.save()
#
#
#
#
##summing topics Last Year    
#datafram = original_df.iloc[:,0]
#df_Topics = datafram.values.tolist()
#datafram1 = original_df.iloc[:,1:]
#df_keywords = datafram1.values.tolist()
#df_keywords_list = sum(df_keywords, [])
##removing nan
#df_keywords_list_only = [i for i in df_keywords_list if i is not np.nan]
#
#df_key_topic = original_df.values.tolist()
#df_key_topic_list = sum(df_key_topic, [])
#df_key_topic_list_only =  [i for i in df_key_topic_list if i is not np.nan]
#
##df22 = df.set_index(df['week'])  
#df_Sum_Topics = df[df_Topics].sum(axis = 0) #yearly sum of topics with scores
#df_only_topics = df[df_Topics]
#
##uncatogorize topics
#df_zeros = df.loc[df[df_Topics].sum(axis=1) == 0]
#
##Normalizing data, everything become 1 except 0
##df3 = df.drop(['month', 'week', 'Loesung', 'Kurzbeschreibung', 'Geoffnet'], axis = 1)
#df2 = df.loc[(df[df_Topics]!=0).any(axis=1)]
##only topics
#df4 = df.loc[(df[df_Topics]!=0).any(axis=1)]
##searching in topics and putting 1 for max and 0 for all other in row
#norm = (df2.eq(df2[df_Topics].max(axis=1), axis=0)).astype(int)       
#
#
#norm_topics = norm[df_Topics]
#df2.drop(df_Topics, axis=1, inplace=True)
#df4.drop(df_key_topic_list_only, axis=1, inplace=True)
#normalize = (df2.join(norm_topics)) 
#normalize_topics = (df4.join(norm_topics)) 
#
#
##creating new column with topic name
#normalize_topics['Assigned Topic'] = normalize_topics[df_Topics].idxmax(axis=1)       
##normalize_topics[df_Topics][normalize_topics[df_Topics].sum(axis=1) > 1]
#
##weekly and monthly frequency of topics and keywords
#weekly = normalize.groupby(['week']).sum().drop(['month'], axis = 1)
##monthly = normalize.groupby(['month']).sum().drop(['week'], axis = 1)
##only topics
#weekly_topics = weekly[df_Topics]
##monthly_topics = monthly[df_Topics]
